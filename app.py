from flask import Flask, render_template, request, jsonify, session, appcontext_popped
import sqlite3
import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import plotly.graph_objs as go
import datetime
import plotly.io as pio
import openpyxl
from flask_oidc import OpenIDConnect
from flask import Flask, redirect, url_for

#imports of other python functions:
from functions.student_services import students_attendance, get_student_dashboard_data
# from functions.teacher_services import (daily_attendance_summary, activity_attendance,
#                                         sport_attendance_by_year, get_teacher_dashboard_data,
#                                         get_sport_popularity, perfect_attendance_students,
#                                         staff_workload_analysis, low_attendance_students, attendance_streak_tracker,
#                                         calculate_student_streaks, get_single_student_attendance, get_available_teams,
#                                         get_team_attendance_data, get_team_summary_stats, get_unique_filter_options,
#                                         get_team_player_stats)

from functions.teacher_services import (daily_attendance_summary, activity_attendance,
    sport_attendance_by_year, get_teacher_dashboard_data, get_sport_popularity,
    perfect_attendance_students, staff_workload_analysis, low_attendance_students,
    attendance_streak_tracker, calculate_student_streaks, get_single_student_attendance,
    get_available_teams, get_team_attendance_data, get_team_summary_stats,
    get_team_player_stats, get_unique_filter_options)





import process_excel_upload
from process_excel_upload import create_connection

app = Flask(__name__)
app.secret_key = "hgytdytdtrds324strcd"

app.config["DATABASE_PATH"] = os.path.join(app.root_path, "dataBase")



#copied auth code for sbhs login page
app.config["OIDC_CLIENT_SECRETS"] = "client_secrets.json"
# This HAS to match the registered scopes
app.config["OIDC_SCOPES"] = "openid profile"
# Prefix the routes added by this library to prevent any
# collisions with our routes.
oidc = OpenIDConnect(app, prefix="/oidc/")


@app.context_processor
def inject_test_user_info(): #Only for testing bcause theres no actual teacher ID testing
    return {
        'is_logged_in': True,
        'user_type': 'teacher',  # Change to 'student' to test student navbar
        'user_name': 'Test Teacher',

    }

# @app.context_processor #renders information which all templates can access as like context before doing anything
# def inject_user_info(): #function enables for different nav bars
#
#     # Check if user is logged in
#     is_logged_in = False
#     user_type = None
#     user_name = None
#     user_id = None
#
#     if oidc.user_loggedin:
#         is_logged_in = True
#         oidc_profile = session.get("oidc_auth_profile")
#
#         if oidc_profile:
#             # Check if it's a student (has student_id)
#             if "student_id" in oidc_profile:
#                 user_type = "student"
#                 user_id = oidc_profile.get('student_id')
#                 user_name = oidc_profile.get('name', 'Student')
#             else:
#                 # It's a teacher/staff
#                 user_type = "teacher"
#                 user_name = oidc_profile.get('name', 'Teacher')
#
#     return {
#         'is_logged_in': is_logged_in,
#         'user_type': user_type,
#         'user_name': user_name,
#         'user_id': user_id
#     }


#Functions NOT related to routing
def parse_date_flexible(date_string):
#    Takes in date string which can be multiple formats
#    and converst them into a datetime object

# all possible date formats to try
    date_formats = [
        '%Y-%m-%d %H:%M:%S',  #Type 1 (Large dataset format): "2024-11-09 14:30:00"
        '%d %b %Y',  #Type 2 (Old realistic data format): "9 Nov 2024"
        '%d-%m-%Y',  #Type3: "09-11-2024"
        '%Y-%m-%d',  #Typ 4 : "2024-11-09"
        '%d/%m/%Y',  #Type 5: "09/11/2024"
        '%m/%d/%Y',  #Type 6: "11/09/2024"
        '%d %B %Y',  #Type 7: "9 November 2024"
    ]

# Try each format until one works
    for dateformat in date_formats:
        try:
            return datetime.datetime.strptime(date_string, dateformat)
        except ValueError: #Keep trying if the datetime doesn't work
            continue

    # if none of them work then send an error
    raise ValueError(f"Could not take in data in format: '{date_string}'. Supported formats: {date_formats}")

def parse_time_flexible(time_string):

#    Does the same thing as date but for time
#    and then returns time object

# all possible time formats
    time_formats = [
        '%I:%M%p',  # Original: "2:30PM"
        '%H:%M',  # 24-hour: "14:30"
        '%I:%M %p',  # With space: "2:30 PM"
        '%H:%M:%S',  # With seconds: "14:30:00"
    ]

# reiterate through all formats till one works
    for timeformat in time_formats:
        try:
            return datetime.datetime.strptime(time_string, timeformat).time()
        except ValueError:
            continue

    # If none work, return a default time
    print(f"System was not able to take in the format of '{time_string}', using default 12:00 PM")
    return datetime.time(12, 0)  # Default to 12:00 PM

def parse_excel_row_flexible(row):
#function basically lets me use both small and large excel sheet as one starts with ID and other starts with Name

    if row[0]:
        first_value = str(row[0])
    else:
        first_value = ""
    if row[1]:
        second_value = str(row[1])
    else:
        second_value = ""


    if first_value.isdigit(): #checks if its a digit, if it is assign these variables
        (
            student_id, full_name, year, boarder, house, homeroom,
            campus, gender, birthdate, secondary, email, team, activity,
            session, date, start_time, end_time, session_staff, attendance,
            for_fixture, flags, cancelled
        ) = row
    else: #if starts with name then assign these variables
        (
            full_name, student_id, year, boarder, house, homeroom,
            campus, gender, birthdate, secondary, email, team, activity,
            session, date, start_time, end_time, session_staff, attendance,
            for_fixture, flags, cancelled
        ) = row


    return ( #hence return the way the code takes it after the assignment
        student_id, full_name, year, boarder, house, homeroom,
        campus, gender, birthdate, secondary, email, team, activity,
        session, date, start_time, end_time, session_staff, attendance,
        for_fixture, flags, cancelled
    )


#Routing:
@app.route('/', methods=['GET'])
def home():
    # Check if user is already logged in
    if oidc.user_loggedin:
        oidc_profile = session.get("oidc_auth_profile")

        if oidc_profile:
            # Check if user is a student
            if "student_id" in oidc_profile:
                return redirect("/student-dashboard")
            else:
                # Assume teacher/staff if not a student
                return redirect("/teacher-dashboard")

    # If not logged in, show the login page
    return render_template('home.html')

@app.route('/logout')
def logout():
    oidc.logout()
    session.clear()
    return redirect("/")

@app.route('/settings')
def settings():
    return render_template('settings.html')

@app.route('/update_settings', methods=['POST'])
def update_settings():
    # Get font choice from form
    font_choice = request.form.get('font_choice', 'default')

    # Store setting in session
    session['font_choice'] = font_choice

    return redirect(url_for('settings'))


@app.route('/teacher-dashboard')
def teacher_dashboard():
    # if not oidc.user_loggedin:
    #     return redirect("/")
    #
    # oidc_profile = session.get("oidc_auth_profile")
    # if "student_id" in oidc_profile:
    #     return redirect("/student-dashboard")  # Wrong user type
    #
    # # Get teacher info from profile
    # teacher_name = oidc_profile.get('name', 'Teacher')
    #
    # # Get real data from database
    dashboard_data = get_teacher_dashboard_data()
    #
    # return render_template('Teacher/teacher-dashboard.html',
    #                        teacher_name=teacher_name,
    #                        **dashboard_data)
    return render_template('Teacher/teacher-dashboard.html', **dashboard_data)


@app.route('/daily-attendance-dashboard', methods=['GET', 'POST'])
def daily_attendance_dashboard():
    results = []
    year_ID = ""

    if request.method == 'POST':
        year_ID = request.form.get('year_ID', '')
        results = daily_attendance_summary(year_ID)  # Call function from external file

    return render_template('Teacher/daily-attendance-dashboard.html', results=results, year_ID=year_ID)
@app.route('/daily-attendance-dashboard-graph', methods=['GET', 'POST'])
def daily_attendance_graph():
    results = []
    graph_html = ""
    year_ID = ""


    if request.method == 'POST':
        # Get the year_ID from the form input
        year_ID = request.form.get('year_ID', '')
        results = daily_attendance_summary(year_ID)  # Fetch data from DB

        if results:
            # Extract data for the graph
            dates = [record[0][:10] for record in results]  # Date column
            present_counts = [record[1] for record in results]  # Present students count
            explained_abs_counts = [record[2] for record in results]  # Explained absences
            unexplained_abs_counts = [record[3] for record in results]  # Unexplained absences

            # Create Plotly Line Graph
            fig = go.Figure()

            fig.add_trace(go.Scatter(x=dates, y=present_counts, mode='lines+markers', name="Present", line=dict(color='green')))
            fig.add_trace(go.Scatter(x=dates, y=explained_abs_counts, mode='lines+markers', name="Explained Absence", line=dict(color='orange')))
            fig.add_trace(go.Scatter(x=dates, y=unexplained_abs_counts, mode='lines+markers', name="Unexplained Absence", line=dict(color='red')))

            # Layout Settings
            fig.update_layout(
                title=f'Daily Attendance Trends for {year_ID}',
                xaxis_title='Date',
                yaxis_title='Number of Students',
                xaxis=dict(type='category')  # Ensures dates display properly
            )

            graph_html = fig.to_html(full_html=False)

    return render_template('Teacher/daily-attendance-dashboard.html', results=results, graph_html=graph_html, year_ID=year_ID)


@app.route('/average-attendance-per-activity', methods=['GET', 'POST'])
def average_attendance_per_activity():
    results = []
    year_ID = ""

    if request.method == 'POST':
        year_ID = request.form.get('year_ID', '')
        results = activity_attendance(year_ID)

    return render_template('Teacher/average-attendance-per-activity.html', results=results, year_ID=year_ID)
@app.route('/average-attendance-per-activity-graph', methods=['GET', 'POST'])
def average_attendance_per_activity_graph():
    results = []
    graph_html = ""
    year_ID = ""

    if request.method == 'POST':
        # Get the year_ID from the form input
        year_ID = request.form.get('year_ID', '')
        # Sanitize input to prevent injection
        year_ID = year_ID.strip()

        results = activity_attendance(year_ID)

        if results:
            activities = [record[0] for record in results]  # Activity names
            attendance_percentages = [record[3] for record in results]  # Average attendance percentages

            # Create Plotly Bar Graph for Average Attendance per Activity
            fig = go.Figure()

            fig.add_trace(go.Bar(x=activities, y=attendance_percentages, name="Average Attendance"))

            # Layout Settings
            fig.update_layout(
                title=f'Average Attendance per Activity for {year_ID}',
                xaxis_title='Activity',
                yaxis_title='Average Attendance (%)',
                xaxis_tickangle=-45  # Rotates the activity names for better visibility
            )

            graph_html = fig.to_html(full_html=False)

    return render_template('Teacher/average-attendance-per-activity.html', results=results, graph_html=graph_html,
                           year_ID=year_ID)


# @app.route('/individual_student_attendance', methods=['GET', 'POST'])
# def individual_student_attendance():
#     results = []
#     graph_html = ""
#     if request.method == 'POST':
#         search_ID = request.form.get('search_ID', '')
#         results = students_attendance(search_ID)
#
#
#
#     # creates attendence graph if we have data
#     if results:
#         #gets attendance data by date
#         dates = [record[4] for record in results] # date - referenced above "for my own reference"
#         statuses = [record[3] for record in results] # attendence
#
#         # makes bar chart data
#         attendance_count = {"Present": 0, "Explained absence": 0, "Unexplained absence": 0}
#         for status in statuses:
#             if status in attendance_count:  # Only count if it's one of the two statuses
#                 attendance_count[status] += 1
#
#         # ploty graph creation
#         data = [
#             go.Bar(
#                 x=list(attendance_count.keys()),
#                 y=list(attendance_count.values()),
#                 marker=dict(color=['green', 'red', 'orange'])
#             )
#         ]
#         #green = present
#         #red = away
#         #orange = late
#
#         #titlting for x and y axis and graph
#         layout = go.Layout(
#             title=f'Attendance Summary for Student {search_ID}', # title
#             xaxis=dict(title='Status'),
#             yaxis=dict(title='Count')
#         )
#
#         fig = go.Figure(data=data, layout=layout)
#         graph_html = fig.to_html(full_html=False)
#
#     return render_template('Teacher/individual_student_attendance.html', results=results, graph_html=graph_html) #converst graph to html graph


# @app.route('/individual_student_attendance', methods=['GET', 'POST'])
# def individual_student_attendance():
#     results = []
#     graph_html = ""
#     search_ID = ""
#
#     # For reference,
#     # record[0] = student_id
#     # record[1] = team
#     # record[2] = activity
#     # record[3] = attendance (Present, Absent, late?)
#     # record[4] = date
#
#     if request.method == 'POST':
#         search_ID = request.form.get('search_ID', '')
#
#         # Incase other page redirects here check.
#         if search_ID:
#             results = students_attendance(search_ID)
#
#             # Create attendance graph if we have data (your existing code)
#             if results:
#                 # Extract attendance statuses
#                 statuses = [record[3] for record in results]
#                 attendance_count = {"Present": 0, "Explained absence": 0, "Unexplained absence": 0}
#
#                 for status in statuses:
#                     if status in attendance_count:
#                         attendance_count[status] += 1
#
#                 # Create Plotly bar chart
#                 data = [
#                     go.Bar(
#                         x=list(attendance_count.keys()),
#                         y=list(attendance_count.values()),
#                         marker=dict(color=['green', 'orange', 'red'])
#                     )
#                 ]
#
#                 layout = go.Layout(
#                     title=f'Attendance Summary for Student {search_ID}',
#                     xaxis=dict(title='Status'),
#                     yaxis=dict(title='Count')
#                 )
#
#                 fig = go.Figure(data=data, layout=layout)
#                 graph_html = fig.to_html(full_html=False)
#
#     # Pass search_ID to template so form shows the searched student ID
#     return render_template('Teacher/individual_student_attendance.html',
#                            results=results,
#                            graph_html=graph_html,
#                            search_ID=search_ID) #todo replace this back


@app.route('/individual_student_attendance', methods=['GET', 'POST'])
def individual_student_attendance():
    results = []
    filtered_results = []
    graph_html = ""
    search_ID = ""
    student_name = ""
    attendance_stats = {}
    unique_activities = []
    date_range_start = ""
    date_range_end = ""
    activity_filter = ""
    status_filter = ""

    if request.method == 'POST':
        search_ID = request.form.get('search_ID', '')
        activity_filter = request.form.get('activity_filter', '')
        status_filter = request.form.get('status_filter', '')

        if search_ID:
            results = students_attendance(search_ID)

            if results:
                # Get student name
                student_name = f"Student {search_ID}"

                # apply filters
                filtered_results = []
                for result in results:
                    should_include = True
                    # if there is an activity filter and its not in results remove it so filter it out
                    if activity_filter and activity_filter not in result[2]:
                        should_include = False

                    # same for status
                    if status_filter and status_filter not in result[3]:
                        should_include = False

                    # if the results are what the filter is just add it tot he filtered resutls
                    if should_include:
                        filtered_results.append(result)

                # calculate attendance statistics
                present_count = 0
                explained_count = 0
                unexplained_count = 0

                for result in results:
                    if result[3] == "Present":
                        present_count += 1
                    elif result[3] == "Explained absence":
                        explained_count += 1
                    elif result[3] == "Unexplained absence":
                        unexplained_count += 1

                total_sessions = len(results)
                if total_sessions > 0:
                    attendance_percentage = round((present_count / total_sessions * 100), 1)
                else:
                    attendance_percentage = 0

                attendance_stats = {
                    'present': present_count,
                    'explained': explained_count,
                    'unexplained': unexplained_count,
                    'percentage': attendance_percentage
                }

                # Get unique activities using basic loops
                unique_activities = []
                for result in results:
                    activity = result[2]
                    if activity not in unique_activities:
                        unique_activities.append(activity)
                unique_activities.sort()

                # Get date range using basic loops
                if results:
                    dates = []
                    for result in results:
                        date_part = result[4][:10]  # Get just the date part
                        dates.append(date_part)

                    date_range_start = min(dates)
                    date_range_end = max(dates)

                # Create Plotly graph
                data = [
                    go.Bar(
                        x=["Present", "Explained absence", "Unexplained absence"],
                        y=[present_count, explained_count, unexplained_count],
                        marker=dict(color=['green', 'orange', 'red'])
                    )
                ]

                layout = go.Layout(
                    title=f'Attendance Summary for {student_name}',
                    xaxis=dict(title='Status'),
                    yaxis=dict(title='Count')
                )

                fig = go.Figure(data=data, layout=layout)
                graph_html = fig.to_html(full_html=False)
            else:
                student_name = "Student Not Found"
                filtered_results = []

    return render_template('Teacher/individual_student_attendance.html',
                           results=results,
                           filtered_results=filtered_results,
                           graph_html=graph_html,
                           search_ID=search_ID,
                           student_name=student_name,
                           attendance_stats=attendance_stats,
                           unique_activities=unique_activities,
                           date_range_start=date_range_start,
                           date_range_end=date_range_end,
                           activity_filter=activity_filter,
                           status_filter=status_filter)


@app.route('/test-student/<student_id>')
def test_student_lookup(student_id): #todo remove this later
    """Test if student lookup is working"""
    results = students_attendance(student_id)

    return {
        'student_id_tested': student_id,
        'results_count': len(results),
        'sample_results': results[:3] if results else [],
        'function_used': 'students_attendance'
    }



@app.route('/sport-popularity', methods=['GET', 'POST'])
def sport_popularity():
    results = []
    year_filter = ""

    if request.method == 'POST':
        year_filter = request.form.get('year_filter', '').strip()  # Changed from year_ID to year_filter

        # Convert to int if provided and not empty, otherwise None
        if year_filter and year_filter != '':
            try:
                year_filter_int = int(year_filter)
            except ValueError:
                year_filter_int = None
        else:
            year_filter_int = None

        results = get_sport_popularity(year_filter_int)

    return render_template('Teacher/sport-popularity.html', results=results,
                           year_filter=year_filter)  # Changed year_ID to year_filter


@app.route('/sport-popularity-graph', methods=['GET', 'POST'])
def sport_popularity_graph():
    results = []
    graph_html = ""
    year_filter = ""

    if request.method == 'POST':
        year_filter = request.form.get('year_filter', '').strip()  # Changed from year_ID to year_filter

        # Convert to int if provided and not empty, otherwise None
        if year_filter and year_filter != '':
            try:
                year_filter_int = int(year_filter)
            except ValueError:
                year_filter_int = None
        else:
            year_filter_int = None

        results = get_sport_popularity(year_filter_int)

        if results:
            sports = [record[0] for record in results]  # Activity names
            enrollments = [record[1] for record in results]  # Total enrolled

            # Create Plotly Bar Graph for Sport Popularity
            fig = go.Figure()

            fig.add_trace(go.Bar(x=sports, y=enrollments, name="Students Enrolled"))

            # Layout Settings
            title_text = f'Sport Popularity for {year_filter}' if year_filter else 'Sport Popularity - All Years'  # Changed year_ID to year_filter
            fig.update_layout(
                title=title_text,
                xaxis_title='Sports/Activities',
                yaxis_title='Number of Students Enrolled',
                xaxis_tickangle=-45
            )

            graph_html = fig.to_html(full_html=False)

    return render_template('Teacher/sport-popularity.html', results=results, graph_html=graph_html,
                           year_filter=year_filter)  # Changed year_ID to year_filter


@app.route('/perfect-attendance', methods=['GET', 'POST'])
def perfect_attendance():
    results = []
    year_ID = ""
    unique_students = 0
    unique_sports = 0

    if request.method == 'POST':
        year_ID = request.form.get('year_ID', '')
        # Convert to int if provided, otherwise None for the function
        year_filter = int(year_ID) if year_ID else None
        results = perfect_attendance_students(year_filter)

        # Calculate summary stats
        if results:
            # Count unique students (some students might be in multiple sports)
            student_ids = []
            sports = []

            for result in results:
                if result[0] not in student_ids:  # result[0] is student_id
                    student_ids.append(result[0])
                if result[3] not in sports:  # result[3] is activity
                    sports.append(result[3])

            unique_students = len(student_ids)
            unique_sports = len(sports)

    return render_template('Teacher/perfect-attendance.html',
                           results=results,
                           year_ID=year_ID,
                           unique_students=unique_students,
                           unique_sports=unique_sports)


@app.route('/staff-workload', methods=['GET', 'POST'])
def staff_workload():
    results = []
    year_filter = ""
    unique_staff_count = 0
    total_sessions = 0
    avg_sessions_per_staff = 0



    if request.method == 'POST':
        year_filter = request.form.get('year_filter', '').strip()

        # Convert to int if provided and not empty, otherwise None
        if year_filter and year_filter != '':
            try:
                year_filter_int = int(year_filter)
            except ValueError:
                year_filter_int = None
        else:
            year_filter_int = None

        results = staff_workload_analysis(year_filter_int)

        if results:
            staff = []

            for result in results:
                staff_name = result[0]
                sessions = result[2]
                if staff_name not in staff:
                    staff.append(staff_name)

                total_sessions += sessions


            unique_staff_count = len(staff)

            if unique_staff_count > 0:
                avg_sessions_per_staff = round(total_sessions / unique_staff_count)
            else:
                avg_sessions_per_staff = 0

    return render_template('Teacher/staff-workload.html', results=results, year_filter=year_filter, unique_staff_count=unique_staff_count, total_sessions=total_sessions, avg_sessions_per_staff=avg_sessions_per_staff)
@app.route('/staff-workload-graph', methods=['POST'])
def staff_workload_graph():
    year_filter = request.form.get('year_filter', '').strip()

    # Convert to int if provided and not empty, otherwise None
    if year_filter and year_filter != '':
        try:
            year_filter_int = int(year_filter)
        except ValueError:
            year_filter_int = None
    else:
        year_filter_int = None

    results = staff_workload_analysis(year_filter_int)
    graph_html = ""

    if results:
        # Group by staff member to get total sessions per staff
        staff_totals = {}
        for result in results:
            staff_name = result[0]  # Staff name
            sessions = result[2]  # Total sessions

            if staff_name in staff_totals:
                staff_totals[staff_name] += sessions
            else:
                staff_totals[staff_name] = sessions

        # Get top 15 staff members by workload
        sorted_staff = sorted(staff_totals.items(), key=lambda x: x[1], reverse=True)[:15]

        staff_names = [item[0] for item in sorted_staff]
        session_counts = [item[1] for item in sorted_staff]

        # Create Plotly Bar Graph
        fig = go.Figure()

        fig.add_trace(go.Bar(
            x=staff_names,
            y=session_counts,
            name="Total Sessions",
            marker=dict(color='lightcoral')
        ))

        # Layout Settings
        title_text = f'Staff Workload - Top 15 Staff Members'
        if year_filter:
            title_text += f' ({year_filter})'

        fig.update_layout(
            title=title_text,
            xaxis_title='Staff Members',
            yaxis_title='Total Sessions Managed',
            xaxis_tickangle=-45
        )

        graph_html = fig.to_html(full_html=False)

    return render_template('Teacher/staff-workload.html',
                           results=results,
                           year_filter=year_filter,
                           graph_html=graph_html)


@app.route('/low-attendance', methods=['GET', 'POST'])
def low_attendance():
    results = []
    year_filter = ""
    attendance_threshold = 85
    total_flagged = 0
    avg_attendance_rate = 0
    lowest_attendance_rate = 100

    if request.method == 'POST':
        year_filter = request.form.get('year_filter', '').strip()
        attendance_threshold = request.form.get('attendance_threshold', '80').strip()

        # Convert inputs
        year_filter_int = None
        if year_filter and year_filter != '':
            try:
                year_filter_int = int(year_filter)
            except ValueError:
                year_filter_int = None

        try:
            attendance_threshold = int(attendance_threshold)
        except ValueError:
            attendance_threshold = 85

        results = low_attendance_students(year_filter_int, attendance_threshold)

        # Calculate summary stats
        if results:
            total_flagged = len(results)
            attendance_rates = [result[8] for result in results]  # attendance_percentage is at index 8
            avg_attendance_rate = round(sum(attendance_rates) / len(attendance_rates), 1)
            lowest_attendance_rate = min(attendance_rates)

    return render_template('Teacher/low-attendance.html',
                           results=results,
                           year_filter=year_filter,
                           attendance_threshold=attendance_threshold,
                           total_flagged=total_flagged,
                           avg_attendance_rate=avg_attendance_rate,
                           lowest_attendance_rate=lowest_attendance_rate)


@app.route('/attendance-streaks', methods=['GET', 'POST'])
def attendance_streaks():
    results = []
    year_filter = ""
    total_students = 0
    longest_streak_ever = 0
    avg_current_streak = 0
    students_with_active_streaks = 0

    if request.method == 'POST':
        year_filter = request.form.get('year_filter', '').strip()

        # Convert to int if provided and not empty, otherwise None
        if year_filter and year_filter != '':
            try:
                year_filter_int = int(year_filter)
            except ValueError:
                year_filter_int = None
        else:
            year_filter_int = None

        # get the streak data
        results = attendance_streak_tracker(year_filter_int)

        #  summary stats
        if results:
            total_students = len(results)
            total_current_streaks = 0 # used for the averages stuff
            students_with_active_streaks = 0
            longest_streak_ever = 0 #total out of all students

            for student in results:
                current_streak = student['current_streak']
                longest_streak = student['longest_streak']

                # Add to totals
                total_current_streaks += current_streak

                # students with at least 1 steak
                if current_streak > 0:
                    students_with_active_streaks += 1

                # Check if this is the longest streak ever
                if longest_streak > longest_streak_ever:
                    longest_streak_ever = longest_streak

            # Calculate average current streak
            if total_students > 0:
                avg_current_streak = round(total_current_streaks / total_students, 1)

            # Sort results by current streak (highest first)
            results.sort(key=lambda x: x['current_streak'], reverse=True)

    return render_template('Teacher/attendance-streaks.html',
                           results=results,
                           year_filter=year_filter,
                           total_students=total_students,
                           longest_streak_ever=longest_streak_ever,
                           avg_current_streak=avg_current_streak,
                           students_with_active_streaks=students_with_active_streaks)


@app.route('/team-attendance', methods=['GET', 'POST'])
def team_attendance():
    teams = []
    year_filter = ""
    semester_filter = ""
    activity_filter = ""

    # Get filter options for dropdowns
    filter_options = get_unique_filter_options()

    if request.method == 'POST':
        year_filter = request.form.get('year_filter', '').strip()
        semester_filter = request.form.get('semester_filter', '').strip()
        activity_filter = request.form.get('activity_filter', '').strip()

        # Convert filters to appropriate types
        if year_filter:
            year_filter_int = int(year_filter)
        else:
            year_filter_int = None
        semester_filter_int = int(semester_filter) if semester_filter else None
        activity_filter_str = activity_filter if activity_filter else None

        teams = get_available_teams(year_filter_int, semester_filter_int, activity_filter_str)

    return render_template('Teacher/team-attendance.html',
                           teams=teams,
                           year_filter=year_filter,
                           semester_filter=semester_filter,
                           activity_filter=activity_filter,
                           filter_options=filter_options)


@app.route('/team-dashboard/<int:team_id>')
def team_dashboard(team_id):

    # Get team summary stats
    team_summary = get_team_summary_stats(team_id)
    if not team_summary:
        return "Team not found", 404

    # Get individual player stats
    player_stats = get_team_player_stats(team_id)

    # Get all attendance records for the team
    attendance_data = get_team_attendance_data(team_id)

    # Create attendance graph
    graph_html = ""
    if attendance_data:
        # Count attendance by status
        present_count = 0
        explained_count = 0
        unexplained_count = 0

        for record in attendance_data:
            status = record[8]  # attendance_status
            if status == "Present":
                present_count += 1
            elif status == "Explained absence":
                explained_count += 1
            elif status == "Unexplained absence":
                unexplained_count += 1

        # Create Plotly graph
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=["Present", "Explained Absence", "Unexplained Absence"],
            y=[present_count, explained_count, unexplained_count],
            marker=dict(color=['green', 'orange', 'red'])
        ))

        fig.update_layout(
            title=f'Team Attendance Overview - {team_summary[0]}',
            xaxis_title='Attendance Status',
            yaxis_title='Number of Sessions',
            height=400
        )

        graph_html = fig.to_html(full_html=False)

    # Get recent sessions (last 10)
    recent_sessions = attendance_data[-10:] if attendance_data else []

    return render_template('Teacher/team-dashboard.html',
                           team_summary=team_summary,
                           player_stats=player_stats,
                           graph_html=graph_html,
                           recent_sessions=recent_sessions,
                           team_id=team_id)


@app.route("/student-login")
def student_login_redirect():
    if oidc.user_loggedin:
        oidc_profile = session.get("oidc_auth_profile")

        # Teachers can potentially log in through the school's OIDC server
        # as well, but we only want students.
        if oidc_profile:
            # Check if user is a student
            if "student_id" in oidc_profile:
                return redirect("/student-dashboard")
            else: #todo NEED TO FIGURE OUT WHICH OIDC RESULT IS TEACHER AND NOT JUST IF NOT STUDENT LOGIC CHANGEEE - MAYBE DELTE
                # Teacher/staff redirect
                return redirect("/teacher-dashboard")
    else:
        # Redirect to  login, then come back to home page
        return oidc.redirect_to_auth_server("/")
@app.route('/student-dashboard')
def student_dashboard():
    # Uncomment these when ready to use authentication
    # if not oidc.user_loggedin:
    #     return redirect("/")
    #
    oidc_profile = session.get("oidc_auth_profile")
    # if "student_id" not in oidc_profile:
    #     return redirect("/teacher-dashboard")  # Wrong user type
    #
    # student_id = oidc_profile.get('student_id', 'Unknown')

    # forceful student id for testing
    student_id = "443194182"  # from the anon excel sheet

    # Get student info from profile
    #student_id = oidc_profile.get('student_id', 'Unknown')

    # Get student dashboard data
    dashboard_data = get_student_dashboard_data(student_id)

    #for streaks
    streak_data = get_single_student_attendance(student_id)


    # Get student's own attendance for graph
    my_attendance_data = students_attendance(student_id)
    my_attendance_graph = ""

    if my_attendance_data:
        # Create attendance graph for this student
        statuses = [record[3] for record in my_attendance_data]
        attendance_count = {"Present": 0, "Explained absence": 0, "Unexplained absence": 0}

        for status in statuses:
            if status in attendance_count:
                attendance_count[status] += 1

        # Create Plotly graph
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=list(attendance_count.keys()),
            y=list(attendance_count.values()),
            marker=dict(color=['green', 'orange', 'red'])
        ))

        fig.update_layout(
            title=f'My Attendance Summary',
            xaxis_title='Status',
            yaxis_title='Count',
            height=400
        )

        my_attendance_graph = fig.to_html(full_html=False)

    # Get recent sessions (last 10)
    recent_sessions = my_attendance_data[-10:] if my_attendance_data else []

    return render_template('Student access/student-dashboard.html',
                           student_id=student_id,
                           my_attendance_graph=my_attendance_graph,
                           recent_sessions=recent_sessions, streak_data=streak_data,
                           **dashboard_data)













#NORMALISED BASE STRUCTURE
# students table:
# - student_id (E.g "443483518")
# - full_name (E.g "John Smith")
# - year_group (E.g "Year 9")
#
# teams table:
# - team_id (auto generated number like 1, 2, 3...)
# - team_name (E.g "14  Team")
# - activity (E.g "Football")
# - year (E.g 2024)
# - semester (like 1 or 2)
# - head_coach (Null cause the data doesn't provde it)
#
# enrollments table (the connector):
# - enrollment_id (auto generated)
# - student_id (connects to students table)
# - team_id (connects to teams table)
#
# attendance_records table:
# - enrollment_id (connects to enrollments table)
# - session_date (E.g "2024-03-15")
# - attendance_status (E.g "Present", "Explained absence", "Unexplained absence")


def normalise_and_insert_data():
    conn = create_connection()
    cursor = conn.cursor()



    #important to order by date so that we read the most recent records first
    cursor.execute('SELECT * FROM staging_full_data '
                   'ORDER BY date DESC')

    rows = cursor.fetchall()



    student_cache = set()
    team_cache = set()
    enrollment_cache = set()
    attendance_cache = set()


    #iterate through each row of data
    for row in rows:
        (
            student_id, full_name, year, boarder, house, homeroom,
            campus, gender, birthdate, secondary, email, team, activity,
            session, date, start_time, end_time, session_staff, attendance,
            for_fixture, flags, cancelled
        ) = parse_excel_row_flexible(row)
        # add all unique students to the student table
        if student_id not in student_cache:
            #write this student into the student table
            cursor.execute('''
                INSERT INTO students (student_id, full_name, 
                year_group, is_boarder, house, homeroom, 
                campus, gender, birth_date, email) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, full_name,
                year, boarder, house, homeroom,
                campus, gender, birthdate, email))

            conn.commit()

            #put the student ID into the cache so that we dont write this student again
            student_cache.add(student_id)

        # add all unique teams to the team table
        team_semester = None
        team_year = None
        date_converted = parse_date_flexible(str(date))
        team_year = date_converted.year
        if date_converted.month <= 6:
            team_semester = 1
        else:
            team_semester = 2

        team_key = (team, activity, team_year, team_semester)
        if team_key not in team_cache:
            # added logic to add teams to the team sql table
            cursor.execute('''
                INSERT INTO teams (team_name, activity, semester, year)
                VALUES (?, ?, ?, ?)
            ''', (team, activity, team_semester, team_year))

            conn.commit()
            team_cache.add(team_key)

        # find team_id
        cursor.execute('''
            SELECT team_id FROM teams
            WHERE team_name = ? AND activity = ? AND semester = ? AND year = ?
        ''', (team, activity, team_semester, team_year))

        team_id = None
        team_row = cursor.fetchone()
        if team_row:
            team_id = team_row[0]

            # create enrollment if not already done
            enrollment_key = (student_id, team_id)
            if enrollment_key not in enrollment_cache:
                cursor.execute('''
                    INSERT INTO enrollments (student_id, team_id)
                    VALUES (?, ?)
                ''', (student_id, team_id))

            conn.commit()
            enrollment_cache.add(enrollment_key)

        # find enrollment_id
        cursor.execute('''
                SELECT enrollment_id FROM enrollments
                WHERE student_id = ? AND team_id = ?
            ''', (student_id, team_id))

        enrollment_row = cursor.fetchone()
        if enrollment_row:
            enrollment_id = enrollment_row[0]

            # create start_datetime by combining date + start_time
            date_part = date_converted.date()


            start_time_part = parse_time_flexible(str(start_time))
            end_time_part = parse_time_flexible(str(end_time))

            start_datetime = datetime.datetime.combine(date_part, start_time_part)
            end_datetime = datetime.datetime.combine(date_part, end_time_part)


            #old datetime code that didnt use the new function which is more flexible
            # create start_datetime by combining date + start_time, might be unneccesary
            # date_part = datetime.datetime.strptime(date, '%d %b %Y').date()  #notorignal -  change back to %Y-%m-%d %H:%M:%S - correct
            # start_time_part = datetime.datetime.strptime(start_time, '%I:%M%p').time()
            # end_time_part = datetime.datetime.strptime(end_time, '%I:%M%p').time()
            # start_datetime = datetime.datetime.combine(date_part, start_time_part)
            # en d_datetime = datetime.datetime.combine(date_part, end_time_part)

            attendance_key = (enrollment_id, start_datetime)
            if attendance_key not in attendance_cache:
                cursor.execute('''
                                INSERT INTO attendance_records (enrollment_id, session_name, session_date, start_time, end_time, staff, attendance_status, is_fixture, has_flags, is_cancelled)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (enrollment_id, session, date, start_datetime, end_datetime, session_staff, attendance, for_fixture, flags, cancelled))

            conn.commit()
            attendance_cache.add(attendance_key)

    #establish which students are enrolled in which teams (enrollment table)

    #capture attendance for each session (attendance_record table)

    conn.close()
@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'GET':
        return render_template('upload.html')

    if request.method == 'POST':

        if 'file' not in request.files:
            return jsonify({'error': 'No file part'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No selected file'})

        if file and file.filename.endswith('.xlsx'):
            # Save a file to a temporary location
            filename = secure_filename(file.filename)   # sanitize filename (disallow special and problematic)
                                                        # characters e.g. ../../file.xlsx)
            file.save(os.path.join('uploads', filename))

            wb = load_workbook(os.path.join('uploads', filename))
            sheet = wb.active

            conn = process_excel_upload.create_connection()
            cursor = conn.cursor()

            # Iterate over the rows (skipping header row)
            cursor.execute('DROP TABLE if exists staging_full_data')
            cursor.execute('DROP TABLE if exists students')
            cursor.execute('DROP TABLE if exists teams')
            cursor.execute('DROP TABLE if exists enrollments')
            cursor.execute('DROP TABLE if exists attendance_records')

            conn.commit()

            process_excel_upload.create_tables()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                (student_id, student_name, year, boarder, house, homeroom, campus, gender, birthdate,
                 secondary, email, team, activity, session, date, start_time, end_time, session_staff, attendance,
                 for_fixture, flags, cancelled) = row

                cursor.execute('''
                        INSERT INTO staging_full_data (
                            student_id, student_name, year, boarder, house, homeroom, campus, gender, birthdate,
                            secondary, email, team, activity, session, date, start_time, end_time, session_staff,
                            attendance, for_fixture, flags, cancelled
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                    student_id, student_name, year, boarder, house, homeroom, campus, gender, birthdate,
                    secondary, email, team, activity, session, date, start_time, end_time, session_staff,
                    attendance, for_fixture, flags, cancelled
                ))

            conn.commit()
            conn.close()

            normalise_and_insert_data()

            return jsonify({"success": "File data added to database"})
        return jsonify({"error": "Invalid file format. Please upload an .xlsx file."})
    return jsonify({"error": "Invalid file format. Please upload an .xlsx file."})



if __name__ == '__main__':
    app.run(debug=True)





















#
#
#
# @app.route('/teacher-year-attendance', methods=['GET', 'POST'])
# def year_attendance():
#     results = []
#     graph_html = ""
#     year_ID = ""
#
#     if request.method == 'POST':
#         # Get the year_ID from the form input
#         year_ID = request.form.get('year_ID', '')
#         results = sport_attendance_by_year(year_ID)
#
#         # Create the attendance graph if we have data
#         if results:
#             # Extract attendance statuses
#             statuses = [record[3] for record in results]  # Attendance status is at index 3
#
#             # Count attendance statuses
#             attendance_count = {"Present": 0, "Explained absence": 0, "Unexplained absence": 0}
#             for status in statuses:
#                 if status in attendance_count:
#                     attendance_count[status] += 1
#
#             # Create Plotly bar chart
#             data = [
#                 go.Bar(
#                     x=list(attendance_count.keys()),
#                     y=list(attendance_count.values()),
#                     marker=dict(color=['green', 'orange', 'red'])
#                 )
#             ]
#
#             # Set graph layout
#             layout = go.Layout(
#                 title=f'Attendance Summary for {year_ID}',
#                 xaxis=dict(title='Attendance Status'),
#                 yaxis=dict(title='Number of Students')
#             )
#
#             # Generate the graph
#             fig = go.Figure(data=data, layout=layout)
#             graph_html = fig.to_html(full_html=False)    # Render the template with results and graph
#     return render_template('daily-attendance-dashboard.html', results=results, graph_html=graph_html, year_ID=year_ID)
#

# @app.route("/student-only-page")
# def student_only_page():
# 	if oidc.user_loggedin:
# 		oidc_profile = session["oidc_auth_profile"]
#
# 		# Teachers can potentially log in through the school's OIDC server
# 		# as well, but we only want students.
# 		if "student_id" not in oidc_profile:
# 			return "SBHS account must be for a student.", 401
#
# 		return f"Hello, {oidc_profile['student_id']}!"
# 	else:
# 		# The argument to this function is what route we want the user to be
# 		# returned to after completing the login. In this case, this page.
# 		return oidc.redirect_to_auth_server("/student-only-page")

#end of copied auth code